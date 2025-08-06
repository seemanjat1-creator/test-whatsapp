import pandas as pd
import openpyxl
import xlrd
from typing import List, Dict, Any, Optional, Tuple
from datetime import datetime
import logging
import re
import numpy as np
from io import BytesIO

logger = logging.getLogger(__name__)

class ExcelProcessor:
    """
    Comprehensive Excel file processor for knowledge base integration.
    Handles multiple worksheets, data types, and creates meaningful text chunks.
    """
    
    def __init__(self):
        self.max_cell_content_length = 1000  # Limit individual cell content
        self.max_rows_per_chunk = 50  # Maximum rows per chunk
        self.preserve_formulas = True  # Whether to include formula text
        self.include_empty_cells = False  # Whether to process empty cells
        
    async def process_excel_file(self, file_path: str, filename: str) -> str:
        """
        Main entry point for Excel file processing.
        Returns concatenated text content from all worksheets.
        """
        try:
            logger.info(f"Starting Excel processing for file: {filename}")
            
            # Determine file type and use appropriate processor
            if filename.lower().endswith('.xlsx'):
                return await self._process_xlsx_file(file_path)
            elif filename.lower().endswith('.xls'):
                return await self._process_xls_file(file_path)
            else:
                raise ValueError(f"Unsupported Excel file format: {filename}")
                
        except Exception as e:
            logger.error(f"Excel processing error for {filename}: {e}")
            raise
    
    async def _process_xlsx_file(self, file_path: str) -> str:
        """Process .xlsx files using openpyxl"""
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=False)
            all_content = []
            
            logger.info(f"Processing {len(workbook.worksheets)} worksheets")
            
            for sheet_idx, worksheet in enumerate(workbook.worksheets):
                sheet_name = worksheet.title
                logger.info(f"Processing worksheet: {sheet_name}")
                
                sheet_content = await self._extract_worksheet_content_openpyxl(
                    worksheet, sheet_name, sheet_idx + 1
                )
                
                if sheet_content.strip():
                    all_content.append(sheet_content)
            
            workbook.close()
            return "\n\n".join(all_content)
            
        except Exception as e:
            logger.error(f"Error processing XLSX file: {e}")
            raise
    
    async def _process_xls_file(self, file_path: str) -> str:
        """Process .xls files using xlrd and pandas"""
        try:
            # Use pandas to read all sheets
            excel_file = pd.ExcelFile(file_path, engine='xlrd')
            all_content = []
            
            logger.info(f"Processing {len(excel_file.sheet_names)} worksheets")
            
            for sheet_idx, sheet_name in enumerate(excel_file.sheet_names):
                logger.info(f"Processing worksheet: {sheet_name}")
                
                try:
                    df = pd.read_excel(excel_file, sheet_name=sheet_name, header=None)
                    sheet_content = await self._extract_worksheet_content_pandas(
                        df, sheet_name, sheet_idx + 1
                    )
                    
                    if sheet_content.strip():
                        all_content.append(sheet_content)
                        
                except Exception as e:
                    logger.warning(f"Error processing sheet {sheet_name}: {e}")
                    continue
            
            excel_file.close()
            return "\n\n".join(all_content)
            
        except Exception as e:
            logger.error(f"Error processing XLS file: {e}")
            raise
    
    async def _extract_worksheet_content_openpyxl(
        self, 
        worksheet: openpyxl.worksheet.worksheet.Worksheet, 
        sheet_name: str, 
        sheet_number: int
    ) -> str:
        """Extract content from worksheet using openpyxl"""
        content_parts = []
        
        # Add worksheet header
        content_parts.append(f"=== WORKSHEET {sheet_number}: {sheet_name} ===")
        
        # Get worksheet dimensions
        max_row = worksheet.max_row
        max_col = worksheet.max_column
        
        if max_row == 1 and max_col == 1:
            # Empty worksheet
            content_parts.append("(Empty worksheet)")
            return "\n".join(content_parts)
        
        logger.info(f"Worksheet dimensions: {max_row} rows x {max_col} columns")
        
        # Process data in chunks to handle large worksheets
        current_chunk = []
        chunk_row_count = 0
        
        # Detect if first row contains headers
        headers = self._detect_headers(worksheet, max_col)
        if headers:
            content_parts.append(f"Headers: {' | '.join(headers)}")
            content_parts.append("")
        
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=1, max_row=max_row), 1):
            row_data = []
            has_content = False
            
            for col_idx, cell in enumerate(row, 1):
                cell_content = self._extract_cell_content(cell)
                
                if cell_content:
                    has_content = True
                    # Include column reference for context
                    col_letter = openpyxl.utils.get_column_letter(col_idx)
                    if headers and col_idx <= len(headers):
                        header = headers[col_idx - 1]
                        row_data.append(f"{header}: {cell_content}")
                    else:
                        row_data.append(f"{col_letter}{row_idx}: {cell_content}")
                elif self.include_empty_cells:
                    row_data.append("")
            
            if has_content:
                current_chunk.append(f"Row {row_idx}: {' | '.join(row_data)}")
                chunk_row_count += 1
                
                # Create chunk when reaching max rows
                if chunk_row_count >= self.max_rows_per_chunk:
                    if current_chunk:
                        content_parts.append("\n".join(current_chunk))
                        content_parts.append("")  # Add spacing between chunks
                    current_chunk = []
                    chunk_row_count = 0
        
        # Add remaining chunk
        if current_chunk:
            content_parts.append("\n".join(current_chunk))
        
        return "\n".join(content_parts)
    
    async def _extract_worksheet_content_pandas(
        self, 
        df: pd.DataFrame, 
        sheet_name: str, 
        sheet_number: int
    ) -> str:
        """Extract content from DataFrame using pandas"""
        content_parts = []
        
        # Add worksheet header
        content_parts.append(f"=== WORKSHEET {sheet_number}: {sheet_name} ===")
        
        if df.empty:
            content_parts.append("(Empty worksheet)")
            return "\n".join(content_parts)
        
        # Replace NaN values with empty strings
        df = df.fillna('')
        
        # Detect headers (first row with mostly non-empty values)
        headers = None
        if not df.iloc[0].isna().all():
            headers = [str(val) for val in df.iloc[0].values if str(val).strip()]
            if len(headers) >= len(df.columns) * 0.5:  # At least 50% non-empty
                content_parts.append(f"Headers: {' | '.join(headers)}")
                content_parts.append("")
                df = df.iloc[1:]  # Skip header row in processing
        
        # Process data in chunks
        current_chunk = []
        chunk_row_count = 0
        
        for idx, row in df.iterrows():
            row_data = []
            has_content = False
            
            for col_idx, value in enumerate(row.values):
                cell_content = self._clean_cell_value(value)
                
                if cell_content:
                    has_content = True
                    if headers and col_idx < len(headers):
                        header = headers[col_idx]
                        row_data.append(f"{header}: {cell_content}")
                    else:
                        col_letter = openpyxl.utils.get_column_letter(col_idx + 1)
                        row_data.append(f"{col_letter}: {cell_content}")
            
            if has_content:
                current_chunk.append(f"Row {idx + 1}: {' | '.join(row_data)}")
                chunk_row_count += 1
                
                # Create chunk when reaching max rows
                if chunk_row_count >= self.max_rows_per_chunk:
                    if current_chunk:
                        content_parts.append("\n".join(current_chunk))
                        content_parts.append("")
                    current_chunk = []
                    chunk_row_count = 0
        
        # Add remaining chunk
        if current_chunk:
            content_parts.append("\n".join(current_chunk))
        
        return "\n".join(content_parts)
    
    def _detect_headers(self, worksheet, max_col: int) -> Optional[List[str]]:
        """Detect if first row contains headers"""
        if worksheet.max_row < 2:
            return None
            
        first_row = list(worksheet.iter_rows(min_row=1, max_row=1, max_col=max_col))[0]
        headers = []
        non_empty_count = 0
        
        for cell in first_row:
            cell_value = self._extract_cell_content(cell)
            if cell_value:
                headers.append(cell_value)
                non_empty_count += 1
            else:
                headers.append(f"Column_{len(headers) + 1}")
        
        # Consider it headers if at least 50% of cells are non-empty
        if non_empty_count >= max_col * 0.5:
            return headers
        
        return None
    
    def _extract_cell_content(self, cell) -> str:
        """Extract meaningful content from a cell"""
        if cell.value is None:
            return ""
        
        content = ""
        
        # Handle different cell value types
        if isinstance(cell.value, str):
            content = cell.value.strip()
        elif isinstance(cell.value, (int, float)):
            # Format numbers appropriately
            if isinstance(cell.value, float) and cell.value.is_integer():
                content = str(int(cell.value))
            else:
                content = str(cell.value)
        elif isinstance(cell.value, datetime):
            content = cell.value.strftime("%Y-%m-%d %H:%M:%S")
        else:
            content = str(cell.value)
        
        # Include formula if available and different from value
        if self.preserve_formulas and hasattr(cell, 'formula') and cell.formula:
            formula = str(cell.formula)
            if formula != content and formula.startswith('='):
                content = f"{content} (Formula: {formula})"
        
        # Limit content length
        if len(content) > self.max_cell_content_length:
            content = content[:self.max_cell_content_length] + "..."
        
        return content
    
    def _clean_cell_value(self, value: Any) -> str:
        """Clean and format cell value from pandas"""
        if pd.isna(value) or value == "":
            return ""
        
        if isinstance(value, str):
            return value.strip()
        elif isinstance(value, (int, float)):
            if isinstance(value, float) and value.is_integer():
                return str(int(value))
            else:
                return str(value)
        elif isinstance(value, datetime):
            return value.strftime("%Y-%m-%d %H:%M:%S")
        else:
            return str(value).strip()
    
    def create_excel_chunks(self, content: str, filename: str) -> List[Dict[str, Any]]:
        """
        Create meaningful chunks from Excel content.
        Maintains worksheet context and logical groupings.
        """
        chunks = []
        
        # Split content by worksheets
        worksheet_sections = content.split("=== WORKSHEET")
        
        for section_idx, section in enumerate(worksheet_sections):
            if not section.strip():
                continue
                
            # Extract worksheet info
            lines = section.strip().split('\n')
            if not lines:
                continue
                
            # Parse worksheet header
            worksheet_info = ""
            content_start_idx = 0
            
            if section_idx > 0:  # Skip first empty section
                worksheet_header = lines[0] if lines else ""
                worksheet_info = f"WORKSHEET {worksheet_header}"
                content_start_idx = 1
            
            # Process worksheet content
            worksheet_content = lines[content_start_idx:]
            
            # Group content into logical chunks
            current_chunk_lines = []
            current_chunk_size = 0
            max_chunk_size = 800  # Characters per chunk
            
            for line in worksheet_content:
                line = line.strip()
                if not line:
                    continue
                
                # Add worksheet context to first chunk
                if not current_chunk_lines and worksheet_info:
                    current_chunk_lines.append(worksheet_info)
                    current_chunk_size += len(worksheet_info)
                
                line_size = len(line)
                
                # Check if adding this line would exceed chunk size
                if current_chunk_size + line_size > max_chunk_size and current_chunk_lines:
                    # Create chunk from current content
                    chunk_content = "\n".join(current_chunk_lines)
                    if chunk_content.strip():
                        chunks.append({
                            "content": chunk_content,
                            "metadata": {
                                "source_file": filename,
                                "worksheet_section": section_idx,
                                "chunk_type": "excel_data",
                                "has_headers": "Headers:" in chunk_content,
                                "row_count": len([l for l in current_chunk_lines if l.startswith("Row")]),
                                "worksheet_info": worksheet_info
                            }
                        })
                    
                    # Start new chunk
                    current_chunk_lines = []
                    current_chunk_size = 0
                    
                    # Add worksheet context to new chunk if needed
                    if worksheet_info:
                        current_chunk_lines.append(worksheet_info)
                        current_chunk_size += len(worksheet_info)
                
                current_chunk_lines.append(line)
                current_chunk_size += line_size
            
            # Add final chunk
            if current_chunk_lines:
                chunk_content = "\n".join(current_chunk_lines)
                if chunk_content.strip():
                    chunks.append({
                        "content": chunk_content,
                        "metadata": {
                            "source_file": filename,
                            "worksheet_section": section_idx,
                            "chunk_type": "excel_data",
                            "has_headers": "Headers:" in chunk_content,
                            "row_count": len([l for l in current_chunk_lines if l.startswith("Row")]),
                            "worksheet_info": worksheet_info
                        }
                    })
        
        logger.info(f"Created {len(chunks)} chunks from Excel file")
        return chunks
    
    def _detect_table_structure(self, worksheet) -> List[Dict[str, Any]]:
        """
        Detect table structures within a worksheet.
        Returns list of table definitions with boundaries and headers.
        """
        tables = []
        
        # Simple table detection: look for header-like patterns
        # This is a basic implementation - can be enhanced for complex scenarios
        
        max_row = min(worksheet.max_row, 1000)  # Limit for performance
        max_col = min(worksheet.max_column, 50)
        
        # Look for potential header rows (rows with mostly text values)
        for row_idx in range(1, min(6, max_row + 1)):  # Check first 5 rows
            row = list(worksheet.iter_rows(min_row=row_idx, max_row=row_idx, max_col=max_col))[0]
            
            text_cells = 0
            non_empty_cells = 0
            
            for cell in row:
                if cell.value is not None:
                    non_empty_cells += 1
                    if isinstance(cell.value, str) and not cell.value.isdigit():
                        text_cells += 1
            
            # If mostly text and mostly non-empty, likely a header
            if (non_empty_cells >= max_col * 0.5 and 
                text_cells >= non_empty_cells * 0.7):
                
                tables.append({
                    "header_row": row_idx,
                    "start_row": row_idx + 1,
                    "end_row": max_row,
                    "start_col": 1,
                    "end_col": max_col,
                    "headers": [self._extract_cell_content(cell) for cell in row]
                })
                break
        
        # If no headers detected, treat entire sheet as one table
        if not tables:
            tables.append({
                "header_row": None,
                "start_row": 1,
                "end_row": max_row,
                "start_col": 1,
                "end_col": max_col,
                "headers": None
            })
        
        return tables
    
    def _format_table_data(
        self, 
        worksheet, 
        table_info: Dict[str, Any]
    ) -> str:
        """Format table data into readable text"""
        content_lines = []
        
        headers = table_info.get("headers")
        start_row = table_info["start_row"]
        end_row = min(table_info["end_row"], start_row + 100)  # Limit rows
        start_col = table_info["start_col"]
        end_col = table_info["end_col"]
        
        # Add headers if available
        if headers:
            content_lines.append(f"Table Headers: {' | '.join(headers)}")
            content_lines.append("")
        
        # Process data rows
        for row_idx in range(start_row, end_row + 1):
            row = list(worksheet.iter_rows(
                min_row=row_idx, max_row=row_idx, 
                min_col=start_col, max_col=end_col
            ))[0]
            
            row_data = []
            has_content = False
            
            for col_idx, cell in enumerate(row):
                cell_content = self._extract_cell_content(cell)
                if cell_content:
                    has_content = True
                    if headers and col_idx < len(headers):
                        row_data.append(f"{headers[col_idx]}: {cell_content}")
                    else:
                        col_letter = openpyxl.utils.get_column_letter(start_col + col_idx)
                        row_data.append(f"{col_letter}: {cell_content}")
            
            if has_content:
                content_lines.append(f"Row {row_idx}: {' | '.join(row_data)}")
        
        return "\n".join(content_lines)
    
    def _extract_merged_cell_info(self, worksheet) -> Dict[str, str]:
        """Extract information about merged cells"""
        merged_info = {}
        
        for merged_range in worksheet.merged_cells.ranges:
            # Get the top-left cell value
            top_left_cell = worksheet[merged_range.min_row][merged_range.min_col - 1]
            if top_left_cell.value:
                range_str = str(merged_range)
                merged_info[range_str] = self._extract_cell_content(top_left_cell)
        
        return merged_info
    
    def validate_excel_file(self, file_path: str, filename: str) -> Tuple[bool, str]:
        """
        Validate Excel file before processing.
        Returns (is_valid, error_message)
        """
        try:
            # Check file extension
            if not filename.lower().endswith(('.xlsx', '.xls')):
                return False, "File must be an Excel file (.xlsx or .xls)"
            
            # Try to open the file
            if filename.lower().endswith('.xlsx'):
                try:
                    workbook = openpyxl.load_workbook(file_path, read_only=True)
                    if len(workbook.worksheets) == 0:
                        return False, "Excel file contains no worksheets"
                    workbook.close()
                except Exception as e:
                    if "password" in str(e).lower():
                        return False, "Password-protected Excel files are not supported"
                    return False, f"Invalid or corrupted Excel file: {str(e)}"
            
            elif filename.lower().endswith('.xls'):
                try:
                    excel_file = pd.ExcelFile(file_path, engine='xlrd')
                    if len(excel_file.sheet_names) == 0:
                        return False, "Excel file contains no worksheets"
                    excel_file.close()
                except Exception as e:
                    if "password" in str(e).lower():
                        return False, "Password-protected Excel files are not supported"
                    return False, f"Invalid or corrupted Excel file: {str(e)}"
            
            return True, ""
            
        except Exception as e:
            logger.error(f"Excel validation error: {e}")
            return False, f"Error validating Excel file: {str(e)}"
    
    def get_excel_metadata(self, file_path: str, filename: str) -> Dict[str, Any]:
        """Extract metadata from Excel file"""
        metadata = {
            "file_type": "excel",
            "worksheets": [],
            "total_sheets": 0,
            "has_formulas": False,
            "has_charts": False,
            "estimated_rows": 0,
            "estimated_cols": 0
        }
        
        try:
            if filename.lower().endswith('.xlsx'):
                workbook = openpyxl.load_workbook(file_path, data_only=False)
                
                for worksheet in workbook.worksheets:
                    sheet_meta = {
                        "name": worksheet.title,
                        "rows": worksheet.max_row,
                        "columns": worksheet.max_column,
                        "has_data": worksheet.max_row > 1 or worksheet.max_column > 1
                    }
                    metadata["worksheets"].append(sheet_meta)
                    metadata["estimated_rows"] += sheet_meta["rows"]
                    metadata["estimated_cols"] = max(metadata["estimated_cols"], sheet_meta["columns"])
                
                metadata["total_sheets"] = len(workbook.worksheets)
                workbook.close()
                
            elif filename.lower().endswith('.xls'):
                excel_file = pd.ExcelFile(file_path, engine='xlrd')
                
                for sheet_name in excel_file.sheet_names:
                    try:
                        df = pd.read_excel(excel_file, sheet_name=sheet_name, header=None)
                        sheet_meta = {
                            "name": sheet_name,
                            "rows": len(df),
                            "columns": len(df.columns),
                            "has_data": not df.empty
                        }
                        metadata["worksheets"].append(sheet_meta)
                        metadata["estimated_rows"] += sheet_meta["rows"]
                        metadata["estimated_cols"] = max(metadata["estimated_cols"], sheet_meta["columns"])
                    except Exception as e:
                        logger.warning(f"Error reading sheet {sheet_name}: {e}")
                
                metadata["total_sheets"] = len(excel_file.sheet_names)
                excel_file.close()
                
        except Exception as e:
            logger.error(f"Error extracting Excel metadata: {e}")
        
        return metadata

# Global instance
excel_processor = ExcelProcessor()