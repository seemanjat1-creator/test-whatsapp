import pytest
import tempfile
import os
from datetime import datetime
from app.services.excel_processor import excel_processor
import openpyxl
import pandas as pd

class TestExcelProcessor:
    """Test suite for Excel file processing functionality"""
    
    @pytest.fixture
    def sample_xlsx_file(self):
        """Create a sample XLSX file for testing"""
        wb = openpyxl.Workbook()
        
        # First worksheet with headers and data
        ws1 = wb.active
        ws1.title = "Sales Data"
        
        # Add headers
        headers = ["Product", "Price", "Quantity", "Total"]
        for col, header in enumerate(headers, 1):
            ws1.cell(row=1, column=col, value=header)
        
        # Add sample data
        data = [
            ["Laptop", 999.99, 5, "=B2*C2"],
            ["Mouse", 25.50, 10, "=B3*C3"],
            ["Keyboard", 75.00, 8, "=B4*C4"]
        ]
        
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws1.cell(row=row_idx, column=col_idx, value=value)
        
        # Second worksheet
        ws2 = wb.create_sheet("Customer Info")
        ws2.cell(row=1, column=1, value="Customer Name")
        ws2.cell(row=1, column=2, value="Email")
        ws2.cell(row=2, column=1, value="John Doe")
        ws2.cell(row=2, column=2, value="john@example.com")
        
        # Save to temporary file
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        wb.save(temp_file.name)
        wb.close()
        
        yield temp_file.name
        
        # Cleanup
        os.unlink(temp_file.name)
    
    @pytest.fixture
    def sample_xls_file(self):
        """Create a sample XLS file for testing"""
        # Create DataFrame
        df = pd.DataFrame({
            'Name': ['Alice', 'Bob', 'Charlie'],
            'Age': [25, 30, 35],
            'City': ['New York', 'London', 'Tokyo']
        })
        
        # Save to temporary file
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xls')
        df.to_excel(temp_file.name, index=False, engine='xlwt')
        
        yield temp_file.name
        
        # Cleanup
        os.unlink(temp_file.name)
    
    @pytest.mark.asyncio
    async def test_xlsx_processing(self, sample_xlsx_file):
        """Test XLSX file processing"""
        content = await excel_processor.process_excel_file(sample_xlsx_file, "test.xlsx")
        
        assert content is not None
        assert len(content) > 0
        assert "WORKSHEET 1: Sales Data" in content
        assert "WORKSHEET 2: Customer Info" in content
        assert "Product" in content  # Header
        assert "Laptop" in content  # Data
        assert "john@example.com" in content  # Email data
    
    @pytest.mark.asyncio
    async def test_xls_processing(self, sample_xls_file):
        """Test XLS file processing"""
        content = await excel_processor.process_excel_file(sample_xls_file, "test.xls")
        
        assert content is not None
        assert len(content) > 0
        assert "Alice" in content
        assert "New York" in content
        assert "Age" in content  # Header
    
    def test_excel_validation_valid_files(self, sample_xlsx_file, sample_xls_file):
        """Test Excel file validation for valid files"""
        # Test XLSX
        is_valid, error = excel_processor.validate_excel_file(sample_xlsx_file, "test.xlsx")
        assert is_valid is True
        assert error == ""
        
        # Test XLS
        is_valid, error = excel_processor.validate_excel_file(sample_xls_file, "test.xls")
        assert is_valid is True
        assert error == ""
    
    def test_excel_validation_invalid_files(self):
        """Test Excel file validation for invalid files"""
        # Test non-Excel file
        is_valid, error = excel_processor.validate_excel_file("test.txt", "test.txt")
        assert is_valid is False
        assert "Excel file" in error
        
        # Test non-existent file
        is_valid, error = excel_processor.validate_excel_file("nonexistent.xlsx", "nonexistent.xlsx")
        assert is_valid is False
    
    def test_excel_chunking(self, sample_xlsx_file):
        """Test Excel content chunking"""
        # First get the content
        import asyncio
        content = asyncio.run(excel_processor.process_excel_file(sample_xlsx_file, "test.xlsx"))
        
        # Create chunks
        chunks = excel_processor.create_excel_chunks(content, "test.xlsx")
        
        assert len(chunks) > 0
        assert all(isinstance(chunk, dict) for chunk in chunks)
        assert all("content" in chunk and "metadata" in chunk for chunk in chunks)
        
        # Check metadata
        for chunk in chunks:
            metadata = chunk["metadata"]
            assert "source_file" in metadata
            assert "chunk_type" in metadata
            assert metadata["chunk_type"] == "excel_data"
            assert "worksheet_section" in metadata
    
    def test_excel_metadata_extraction(self, sample_xlsx_file):
        """Test Excel metadata extraction"""
        metadata = excel_processor.get_excel_metadata(sample_xlsx_file, "test.xlsx")
        
        assert metadata["file_type"] == "excel"
        assert metadata["total_sheets"] == 2
        assert len(metadata["worksheets"]) == 2
        assert metadata["worksheets"][0]["name"] == "Sales Data"
        assert metadata["worksheets"][1]["name"] == "Customer Info"
        assert metadata["estimated_rows"] > 0
        assert metadata["estimated_cols"] > 0
    
    def test_cell_content_extraction(self):
        """Test individual cell content extraction"""
        # Create a test workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # Test different data types
        ws.cell(row=1, column=1, value="Text Content")
        ws.cell(row=1, column=2, value=123.45)
        ws.cell(row=1, column=3, value=datetime(2024, 1, 15))
        ws.cell(row=1, column=4, value="=A1&B1")  # Formula
        
        # Test extraction
        text_content = excel_processor._extract_cell_content(ws.cell(row=1, column=1))
        assert text_content == "Text Content"
        
        number_content = excel_processor._extract_cell_content(ws.cell(row=1, column=2))
        assert "123.45" in number_content
        
        date_content = excel_processor._extract_cell_content(ws.cell(row=1, column=3))
        assert "2024-01-15" in date_content
        
        wb.close()

if __name__ == "__main__":
    # Run basic tests
    import asyncio
    
    async def run_basic_test():
        """Run a basic test to verify Excel processing works"""
        # Create a simple test file
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Test Sheet"
        
        # Add test data
        ws.cell(row=1, column=1, value="Name")
        ws.cell(row=1, column=2, value="Value")
        ws.cell(row=2, column=1, value="Test Item")
        ws.cell(row=2, column=2, value=42)
        
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        wb.save(temp_file.name)
        wb.close()
        
        try:
            # Test processing
            content = await excel_processor.process_excel_file(temp_file.name, "test.xlsx")
            print("Excel processing successful!")
            print(f"Content length: {len(content)}")
            print(f"Content preview: {content[:200]}...")
            
            # Test chunking
            chunks = excel_processor.create_excel_chunks(content, "test.xlsx")
            print(f"Created {len(chunks)} chunks")
            
            return True
            
        except Exception as e:
            print(f"Excel processing failed: {e}")
            return False
        finally:
            os.unlink(temp_file.name)
    
    # Run the test
    success = asyncio.run(run_basic_test())
    if success:
        print("✅ Excel processor is working correctly!")
    else:
        print("❌ Excel processor test failed!")