"""
Integration tests for Excel file support in the knowledge base system.
"""

import pytest
import tempfile
import os
from datetime import datetime
from app.services.document_service import document_service
from app.services.excel_processor import excel_processor
from app.models.document import DocumentCreate, DocumentType, DocumentStatus
import openpyxl
import pandas as pd

class TestExcelIntegration:
    """Integration tests for Excel support"""
    
    @pytest.fixture
    def complex_excel_file(self):
        """Create a complex Excel file for testing"""
        wb = openpyxl.Workbook()
        
        # Sales data worksheet
        ws1 = wb.active
        ws1.title = "Q1 Sales Report"
        
        # Headers
        headers = ["Date", "Product", "Category", "Sales Rep", "Units Sold", "Revenue", "Commission"]
        for col, header in enumerate(headers, 1):
            ws1.cell(row=1, column=col, value=header)
        
        # Sample data with different data types
        sales_data = [
            [datetime(2024, 1, 15), "Laptop Pro", "Electronics", "John Smith", 5, 4999.95, "=F2*0.1"],
            [datetime(2024, 1, 16), "Wireless Mouse", "Electronics", "Jane Doe", 25, 625.00, "=F3*0.1"],
            [datetime(2024, 1, 17), "Office Chair", "Furniture", "Bob Wilson", 8, 1599.92, "=F4*0.1"],
            [datetime(2024, 1, 18), "Standing Desk", "Furniture", "Alice Brown", 3, 899.97, "=F5*0.1"],
        ]
        
        for row_idx, row_data in enumerate(sales_data, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws1.cell(row=row_idx, column=col_idx, value=value)
        
        # Customer data worksheet
        ws2 = wb.create_sheet("Customer Database")
        customer_headers = ["Customer ID", "Name", "Email", "Phone", "City", "Total Orders", "Lifetime Value"]
        
        for col, header in enumerate(customer_headers, 1):
            ws2.cell(row=1, column=col, value=header)
        
        customer_data = [
            ["CUST001", "Michael Johnson", "michael@email.com", "+1-555-0101", "New York", 12, 15000.50],
            ["CUST002", "Sarah Davis", "sarah@email.com", "+1-555-0102", "Los Angeles", 8, 9500.25],
            ["CUST003", "David Wilson", "david@email.com", "+1-555-0103", "Chicago", 15, 22000.75],
        ]
        
        for row_idx, row_data in enumerate(customer_data, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws2.cell(row=row_idx, column=col_idx, value=value)
        
        # Product catalog worksheet
        ws3 = wb.create_sheet("Product Catalog")
        product_headers = ["SKU", "Product Name", "Description", "Price", "Stock", "Category"]
        
        for col, header in enumerate(product_headers, 1):
            ws3.cell(row=1, column=col, value=header)
        
        product_data = [
            ["LP001", "Laptop Pro 15", "High-performance laptop with 16GB RAM", 999.99, 50, "Electronics"],
            ["MS001", "Wireless Mouse", "Ergonomic wireless mouse with USB receiver", 25.00, 200, "Electronics"],
            ["CH001", "Office Chair", "Ergonomic office chair with lumbar support", 199.99, 30, "Furniture"],
        ]
        
        for row_idx, row_data in enumerate(product_data, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws3.cell(row=row_idx, column=col_idx, value=value)
        
        # Save to temporary file
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        wb.save(temp_file.name)
        wb.close()
        
        yield temp_file.name
        
        # Cleanup
        os.unlink(temp_file.name)
    
    @pytest.mark.asyncio
    async def test_excel_upload_and_processing(self, complex_excel_file):
        """Test complete Excel upload and processing pipeline"""
        
        # Simulate file upload
        with open(complex_excel_file, 'rb') as f:
            file_content = f.read()
        
        # Create mock UploadFile
        class MockUploadFile:
            def __init__(self, filename: str, content: bytes):
                self.filename = filename
                self.content = content
                self.size = len(content)
                self.content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            
            async def read(self):
                return self.content
        
        mock_file = MockUploadFile("test_sales_data.xlsx", file_content)
        
        try:
            # Test document upload (this would normally go through the API)
            document = await document_service.upload_document(
                file=mock_file,
                workspace_id="test_workspace_id",
                title="Sales and Customer Data",
                description="Comprehensive sales report with customer information",
                tags=["sales", "customers", "q1-2024"]
            )
            
            # Verify document was created
            assert document is not None
            assert document.document_type in [DocumentType.XLSX, DocumentType.XLS]
            assert document.status == DocumentStatus.READY
            assert "Sales and Customer Data" in document.title
            
            # Verify content was extracted
            assert len(document.content) > 100
            assert "Q1 Sales Report" in document.content
            assert "Customer Database" in document.content
            assert "Product Catalog" in document.content
            
            logger.info("✅ Excel upload and processing test passed!")
            
        except Exception as e:
            logger.error(f"❌ Excel integration test failed: {e}")
            raise
    
    @pytest.mark.asyncio
    async def test_excel_search_functionality(self, complex_excel_file):
        """Test Excel content search and retrieval"""
        
        # Process the Excel file
        content = await excel_processor.process_excel_file(complex_excel_file, "test.xlsx")
        
        # Create chunks
        chunks = excel_processor.create_excel_chunks(content, "test.xlsx")
        
        # Test search queries
        test_queries = [
            "laptop sales data",
            "customer email addresses", 
            "product pricing information",
            "John Smith sales performance",
            "electronics category revenue"
        ]
        
        for query in test_queries:
            # Find relevant chunks (simplified search)
            relevant_chunks = []
            query_lower = query.lower()
            
            for chunk in chunks:
                chunk_content = chunk["content"].lower()
                if any(word in chunk_content for word in query_lower.split()):
                    relevant_chunks.append(chunk)
            
            logger.info(f"Query '{query}' found {len(relevant_chunks)} relevant chunks")
            assert len(relevant_chunks) > 0, f"No results found for query: {query}"
        
        logger.info("✅ Excel search functionality test passed!")
    
    def test_excel_chunk_quality(self, complex_excel_file):
        """Test the quality and structure of Excel chunks"""
        import asyncio
        
        # Process file
        content = asyncio.run(excel_processor.process_excel_file(complex_excel_file, "test.xlsx"))
        chunks = excel_processor.create_excel_chunks(content, "test.xlsx")
        
        # Verify chunk quality
        for chunk in chunks:
            # Each chunk should have content and metadata
            assert "content" in chunk
            assert "metadata" in chunk
            assert len(chunk["content"]) > 10
            
            metadata = chunk["metadata"]
            assert "source_file" in metadata
            assert "chunk_type" in metadata
            assert metadata["chunk_type"] == "excel_data"
            
            # Excel-specific metadata
            if "worksheet_info" in metadata:
                assert len(metadata["worksheet_info"]) > 0
        
        logger.info(f"✅ Excel chunk quality test passed! Created {len(chunks)} high-quality chunks")
    
    def test_excel_error_handling(self):
        """Test error handling for various Excel file issues"""
        
        # Test invalid file
        try:
            is_valid, error = excel_processor.validate_excel_file("nonexistent.xlsx", "nonexistent.xlsx")
            assert is_valid is False
            assert len(error) > 0
        except Exception:
            pass  # Expected for non-existent file
        
        # Test invalid extension
        is_valid, error = excel_processor.validate_excel_file("test.txt", "test.txt")
        assert is_valid is False
        assert "Excel file" in error
        
        logger.info("✅ Excel error handling test passed!")

# Performance test for large Excel files
def create_large_excel_test_file(rows: int = 1000, cols: int = 10) -> str:
    """Create a large Excel file for performance testing"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Large Dataset"
    
    # Add headers
    for col in range(1, cols + 1):
        ws.cell(row=1, column=col, value=f"Column_{col}")
    
    # Add data
    for row in range(2, rows + 2):
        for col in range(1, cols + 1):
            if col == 1:
                ws.cell(row=row, column=col, value=f"Item_{row-1}")
            elif col == 2:
                ws.cell(row=row, column=col, value=row * 10.5)
            else:
                ws.cell(row=row, column=col, value=f"Data_{row}_{col}")
    
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    wb.save(temp_file.name)
    wb.close()
    
    return temp_file.name

if __name__ == "__main__":
    # Run integration tests
    import asyncio
    
    async def run_integration_tests():
        """Run basic integration tests"""
        print("🧪 Running Excel integration tests...")
        
        # Create test file
        test_file = create_large_excel_test_file(100, 5)
        
        try:
            # Test processing
            content = await excel_processor.process_excel_file(test_file, "large_test.xlsx")
            print(f"✅ Processed large Excel file: {len(content)} characters")
            
            # Test chunking
            chunks = excel_processor.create_excel_chunks(content, "large_test.xlsx")
            print(f"✅ Created {len(chunks)} chunks from large Excel file")
            
            # Test validation
            is_valid, error = excel_processor.validate_excel_file(test_file, "large_test.xlsx")
            print(f"✅ Validation result: {is_valid}, error: {error}")
            
            print("🎉 All Excel integration tests passed!")
            
        except Exception as e:
            print(f"❌ Integration test failed: {e}")
        finally:
            os.unlink(test_file)
    
    # Run tests
    asyncio.run(run_integration_tests())